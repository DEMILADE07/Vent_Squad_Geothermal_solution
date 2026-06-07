"""Generate data/processed/LCOE_hybrid.xlsx — the extended hybrid workbook.

The stock TNO LCOE.xlsx handles heat only (with power as an alternative).
Challenge 2 asks for the *hybrid* heating + cooling design, so rather than hack
formulas into the original (which stays untouched as the single source of truth),
this writes a clean, self-contained workbook driven by src/lcoe.py:

  Cover          what this file is, provenance, the reference-reproduction check.
  Inputs         every heat- and cooling-side assumption with units and source.
  Outputs        headline LCOE heat / cool / blended, capex, MWth, P50 resource.
  Design_A_vs_B  side-by-side of the two surface designs.
  CF_heat        annual after-tax cash flow for the heat product (auditable).
  CF_cool        annual after-tax cash flow for the cooling product (Design A).

Because it is generated from code, the workbook always agrees with the notebook
and the pipeline — there is no separate spreadsheet to drift.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.lcoe import (
    CoolingCase,
    HeatCase,
    REFERENCE_LCOE_EUR_GJ,
    cooling_economics,
    heat_economics,
    validate_reference,
)
from src.paths import LCOE_HYBRID
from src.surface import SchemeConfig, size_design_a, size_design_b

_HDR = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
_TITLE = Font(bold=True, size=14)
_BOLD = Font(bold=True)
_MONEY = "#,##0.00"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row, labels):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=lab)
        c.font = _HDR
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left")


def build(path=LCOE_HYBRID) -> str:
    cfg = SchemeConfig()
    a = size_design_a(cfg)
    b = size_design_b(cfg)
    heat = a["_heat"]
    cool = a["_cool"]
    ref = validate_reference()

    wb = openpyxl.Workbook()

    # --- Cover -------------------------------------------------------------
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "Hybrid Geothermal Heating + Cooling — Extended LCOE"
    ws["A1"].font = _TITLE
    rows = [
        ("", ""),
        ("Project", "Utrecht district geothermal H&C — 2-doublet Rotliegend scheme"),
        ("Reservoir", "Lower Permian Rotliegend, anchored on BLT-01 (ThermoGIS P50)"),
        ("Source workbook", "TNO LCOE.xlsx (van Wees et al., 2011, v2012.1) — heat-only"),
        ("This file", "Hybrid extension generated from src/lcoe.py (heat + cooling)"),
        ("", ""),
        ("Validation", "TNO heat reference reproduced by our engine:"),
        ("  our engine", f"{ref:.4f} EUR/GJ"),
        ("  TNO published", f"{REFERENCE_LCOE_EUR_GJ:.4f} EUR/GJ"),
        ("  match", "PASS" if abs(ref - REFERENCE_LCOE_EUR_GJ) < 1e-3 else "FAIL"),
        ("", ""),
        ("Headline (Design A, recommended)", ""),
        ("  LCOE heat", f"{a['lcoe_heat_eur_gj']:.2f} EUR/GJ"),
        ("  LCOE cooling", f"{a['lcoe_cool_eur_gj']:.2f} EUR/GJ"),
        ("  Heat delivered (P50)", f"{a['heat_mwth']:.2f} MWth"),
        ("  Cooling delivered", f"{a['cooling_mwth']:.2f} MWth"),
        ("  Total capex", f"{a['capex_total_mln']:.2f} mln EUR"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = _BOLD
        ws.cell(row=i, column=2, value=v)
    _autosize(ws, [34, 60])

    # --- Inputs ------------------------------------------------------------
    ws = wb.create_sheet("Inputs")
    _header_row(ws, 1, ["Block", "Parameter", "Value", "Unit", "Source / note"])
    h: HeatCase = cfg.heat_case()
    k: CoolingCase = CoolingCase(cooling_mwth=cfg.demand_cooling_mwth, ates_pairs=a["ates_pairs"])
    fin = h.fin
    inputs = [
        ("Resource", "Heat delivered (demand-capped)", round(cfg.heat_delivered_mwth, 2), "MWth", "10 MWth demand; resource P50 13.2 MWth"),
        ("Resource", "Production temperature", cfg.t_prod_c, "C", "ThermoGIS P50, BLT-01"),
        ("Resource", "Reinjection temperature", cfg.t_reinject_c, "C", "ThermoGIS standard"),
        ("Heat", "Number of wells", cfg.n_wells, "-", "2 doublets x 2 wells"),
        ("Heat", "Along-hole depth / well", h.ah_depth_m, "m", "deviated NL Rotliegend doublet"),
        ("Heat", "Heat load hours", h.heat_loadhours, "h/yr", "TNO direct-heat default"),
        ("Heat", "Well cost scaling", h.well_cost_scaling, "-", "TNO well-cost factor"),
        ("Heat", "Pump cost", h.pump_cost_mln, "mln EUR", "TNO; workover every 5 yr"),
        ("Heat", "Heat plant invest", h.heat_plant_invest_keur_mwth, "kEUR/MWth", "TNO"),
        ("Heat", "Electricity price (pumps)", h.electricity_price_eur_mwh, "EUR/MWh", "TNO default"),
        ("Heat", "Pump COP", h.pump_cop, "-", "TNO"),
        ("Cooling", "Cooling delivered", k.cooling_mwth, "MWth", "district demand"),
        ("Cooling", "Cooling load hours", k.cooling_loadhours, "h/yr", "summer-peaky, FLEQ"),
        ("Cooling", "ATES well pairs", k.ates_pairs, "-", "warm/cold doublets"),
        ("Cooling", "ATES capex / pair", k.ates_capex_mln_per_pair, "mln EUR", "NL range 0.5-0.8"),
        ("Cooling", "ATES throughput / pair", k.ates_throughput_mwth_per_pair, "MWth", "conservative central of 0.5-2.0 (Fleuchaus 2018)"),
        ("Cooling", "System cooling COP", k.system_cop_cooling, "-", "ATES circulation + HP trim, ~70/30 blend"),
        ("Cooling", "Chiller/HP capex", k.chiller_capex_keur_mwth, "kEUR/MWth", "trim heat pump"),
        ("Finance", "Loan rate", fin.loan_rate, "-", "TNO"),
        ("Finance", "Required equity return", fin.equity_return, "-", "TNO; = discount rate"),
        ("Finance", "Equity / debt share", f"{fin.equity_share:g} / {fin.debt_share:g}", "-", "TNO"),
        ("Finance", "Tax rate", fin.tax, "-", "TNO"),
        ("Finance", "Term loan / depreciation", f"{fin.term_loan_yr} / {fin.depreciation_yr}", "yr", "TNO"),
    ]
    for i, row in enumerate(inputs, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    _autosize(ws, [12, 26, 12, 12, 40])

    # --- Outputs -----------------------------------------------------------
    ws = wb.create_sheet("Outputs")
    _header_row(ws, 1, ["Metric", "Value", "Unit"])
    outs = [
        ("LCOE heat (Design A)", round(a["lcoe_heat_eur_gj"], 2), "EUR/GJ"),
        ("LCOE heat", round(a["lcoe_heat_eur_gj"] * 3.6, 1), "EUR/MWhth"),
        ("LCOE cooling (Design A)", round(a["lcoe_cool_eur_gj"], 2), "EUR/GJ"),
        ("Blended system LCOE", round(
            (a["lcoe_heat_eur_gj"] * heat["npv_energy_gj"]
             + a["lcoe_cool_eur_gj"] * cool["npv_energy_gj"])
            / (heat["npv_energy_gj"] + cool["npv_energy_gj"]), 2), "EUR/GJ"),
        ("Heat delivered (P50)", round(a["heat_mwth"], 2), "MWth"),
        ("Cooling delivered", round(a["cooling_mwth"], 2), "MWth"),
        ("Subsurface capex", round(heat["subsurface_capex_mln"], 2), "mln EUR"),
        ("Heat plant capex", round(heat["heat_capex_mln"], 2), "mln EUR"),
        ("Cooling capex (ATES + chiller)", round(cool["capex_mln"], 2), "mln EUR"),
        ("Total capex", round(a["capex_total_mln"], 2), "mln EUR"),
        ("Implied scheme flow", round(heat["flow_ls"], 1), "L/s"),
        ("TNO heat-only reference", round(REFERENCE_LCOE_EUR_GJ, 2), "EUR/GJ"),
    ]
    for i, row in enumerate(outs, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            if j == 2 and isinstance(val, float):
                c.number_format = _MONEY
    _autosize(ws, [32, 14, 12])

    # --- Design A vs B -----------------------------------------------------
    ws = wb.create_sheet("Design_A_vs_B")
    _header_row(ws, 1, ["Metric", "Design A (ATES+HP)", "Design B (absorption)", "Unit"])
    comp = [
        ("Heat delivered", a["heat_mwth"], b["heat_mwth"], "MWth"),
        ("Cooling delivered", a["cooling_mwth"], b["cooling_mwth"], "MWth"),
        ("Geo heat consumed for cooling", a["geo_heat_consumed_for_cooling_mwth"],
         b["geo_heat_consumed_for_cooling_mwth"], "MWth"),
        ("ATES well pairs", a["ates_pairs"], b["ates_pairs"], "-"),
        ("LCOE heat", a["lcoe_heat_eur_gj"], b["lcoe_heat_eur_gj"], "EUR/GJ"),
        ("LCOE cooling", a["lcoe_cool_eur_gj"], b["lcoe_cool_eur_gj"], "EUR/GJ"),
        ("Total capex", a["capex_total_mln"], b["capex_total_mln"], "mln EUR"),
        ("Cooling electricity", a["cooling_elec_mwh_yr"], b["cooling_elec_mwh_yr"], "MWh/yr"),
    ]
    for i, (m, va, vb, u) in enumerate(comp, start=2):
        ws.cell(row=i, column=1, value=m)
        ca = ws.cell(row=i, column=2, value=round(va, 2) if isinstance(va, float) else va)
        cb = ws.cell(row=i, column=3, value=round(vb, 2) if isinstance(vb, float) else vb)
        ws.cell(row=i, column=4, value=u)
        if isinstance(va, float):
            ca.number_format = _MONEY
            cb.number_format = _MONEY
    note_row = len(comp) + 3
    ws.cell(row=note_row, column=1,
            value="Recommendation: Design A. Lower cooling LCOE; cooling is "
                  "decoupled from heat sales; and a LiBr/H2O absorption chiller "
                  "wants 85-95 C drive heat, above our 77 C reservoir.").font = _BOLD
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    _autosize(ws, [32, 20, 22, 10])

    # --- Cash-flow sheets --------------------------------------------------
    for sheet_name, econ in (("CF_heat", heat), ("CF_cool", cool)):
        ws = wb.create_sheet(sheet_name)
        cf = econ["cashflow"]
        _header_row(ws, 1, ["Year", "Energy (GJ)", "Gross rev (EUR)",
                            "Depreciation (EUR)", "Interest (EUR)",
                            "Loan charge (EUR)", "Tax credit (EUR)",
                            "Net rev after tax (EUR)"])
        keys = ["year", "energy_gj", "gross_rev_eur", "depreciation_eur",
                "interest_eur", "loan_charge_eur", "tax_credit_eur", "net_rev_eur"]
        for r in range(len(cf["year"])):
            for j, key in enumerate(keys, start=1):
                val = float(cf[key][r])
                c = ws.cell(row=r + 2, column=j, value=round(val, 2))
                if j > 1:
                    c.number_format = _MONEY
        _autosize(ws, [6] + [16] * 7)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return str(path)


if __name__ == "__main__":
    print("wrote", build())
